"""DDC Calculateur.xlsx — 2 tableaux côte à côte (avec / sans FY)."""

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from openpyxl.worksheet.table import Table, TableStyleInfo

OUTPUT = r"C:\Users\Sylvain Rio\Downloads\DDC - Calculateur v7.xlsx"
MAX_NODES = 20
BLOCK = 12
DATA_START, DATA_END = 5, 204

SYN_TITLE = 3
SYN_HDR = 4
SYN_DATA = 5
SYN_LAST = SYN_DATA + MAX_NODES * BLOCK - 1

SPACER_COLS = (7, 8)  # G, H

FONT = "Aptos"
HDR = PatternFill("solid", fgColor="1F4E79")
HF = Font(name=FONT, bold=True, color="FFFFFF", size=11)
BF = Font(name=FONT, size=11)
TF = Font(name=FONT, bold=True, size=11, color="1F4E79")
NF = Font(name=FONT, bold=True, size=11, color="1F4E79")
SF = Font(name=FONT, bold=True, size=11, color="1F4E79")
INP = PatternFill("solid", fgColor="FFF2CC")
BD = Border(*(Side(style="thin", color="B4B4B4"),) * 4)
CE = Alignment(horizontal="center", vertical="center")

P = {
    "VGD": '"*Vent G/D*"',
    "VDG": '"*Vent D/G*"',
    "VAA": '"*Vent Av./Arr.*"',
    "VAV": '"*Vent Arr./Av.*"',
}

ROWS = [
    ("Couverture", "c"),
    ("Vent G/D", "xd", "VGD"),
    ("Vent G/D", "xf", "VGD"),
    ("Vent D/G", "xd", "VDG"),
    ("Vent D/G", "xf", "VDG"),
    ("Vent Av./Arr.", "xd", "VAA"),
    ("Vent Av./Arr.", "xf", "VAA"),
    ("Vent Arr./Av.", "xd", "VAV"),
    ("Vent Arr./Av.", "xf", "VAV"),
    ("Neige normale", "n"),
    ("Neige accidentelle", "a"),
]

EXEMPLE_FY = [
    ("47/1", "G : Couverture : PA 6 BS 10 PV15", -116, 7, -1285),
    ("47/2", "Vent G/D dep.(-) Portique 2", 905, 40, 1422),
    ("47/3", "Vent G/D sur.(+) Portique 2", 900, 32, 1444),
    ("47/4", "Vent D/G dep.(-) Cpe - Portique 2", -345, 5, 462),
    ("47/5", "Vent D/G dep.(-) Cpe + Portique 2", -513, -11, -281),
    ("47/6", "Vent D/G sur.(+) Cpe - Portique 2", -349, -2, 485),
    ("47/7", "Vent D/G sur.(+) Cpe + Portique 2", -518, -22, -258),
    ("47/8", "Vent Av./Arr. dep.(-) Portique 2", -241, 205, 815),
    ("47/9", "Vent Av./Arr. sur.(+) Portique 2", -198, 195, 598),
    ("47/10", "Vent Arr./Av. dep.(-) Portique 2", -105, 0, 400),
    ("47/11", "Vent Arr./Av. sur.(+) Portique 2", -37, -11, 56),
    ("47/12", "Neige cas I", -113, 0, -1010),
    ("47/13", "Neige cas II", -56, 0, -505),
    ("47/14", "Neige accidentel", -251, 0, -2244),
]


@dataclass(frozen=True)
class TableCfg:
    name: str
    has_fy: bool
    cas_col: str
    nom_col: str
    fx_col: str
    fy_col: str | None
    fz_col: str
    node_col: str
    syn_col: int
    syn_node_col: str
    syn_fz_col: str
    table_hdr_row: int
    table_last_col: str

    @property
    def b(self):
        return rng(self.nom_col)

    @property
    def fx(self):
        return rng(self.fx_col)

    @property
    def fy(self):
        return rng(self.fy_col) if self.fy_col else None

    @property
    def fz(self):
        return rng(self.fz_col)

    @property
    def n(self):
        return rng(self.node_col)

    @property
    def n_anchor(self):
        return f"SAISIE!${self.node_col}${DATA_START}"

    @property
    def cas_range(self):
        return rng(self.cas_col)

    @property
    def syn_width(self):
        return 5 if self.has_fy else 4


def rng(col):
    return f"SAISIE!${col}${DATA_START}:${col}${DATA_END}"


CFG_FY = TableCfg(
    "AVEC FY",
    True,
    "A", "B", "C", "D", "E", "F",
    1, "F", "D",
    4, "F",
)
CFG_NOFY = TableCfg(
    "SANS FY",
    False,
    "I", "J", "K", None, "L", "M",
    9, "M", "K",
    4, "M",
)


def node_extract_formula(row, cas_col):
    return (
        f'=IF({cas_col}{row}="","",'
        f'VALUE(LEFT({cas_col}{row},FIND("/",{cas_col}{row}&"/")-1)))'
    )


def wrap(z, inner):
    return f'=IF({z}="","",{inner.lstrip("=")})'


def wrap_visible(z, cond, inner):
    return f'=IF(OR({z}="",NOT({cond})),"",{inner.lstrip("=")})'


def family_has_cases(cfg, p, z):
    return f"COUNTIFS({cfg.b},{p},{cfg.n},{z})>0"


def family_has_uplift(cfg, p, z):
    return f"MAXIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z})>0"


def note_dynamic(row, z, fz_col):
    fz = f"{fz_col}{row}"
    return (
        f'=IF({z}="","",IF(OR({fz}="",{fz}="-"),"",'
        f'IF({fz}>0,"(soulevement)",IF({fz}<0,"(compression)","-"))))'
    )


def node_formula(cfg, row, node_col):
    prev_end = row - 1
    prev = f"${node_col}$4:{node_col}{prev_end}" if row > SYN_DATA else f"${node_col}$4:${node_col}{SYN_DATA - 1}"
    return (
        f"=IFERROR(INDEX({cfg.n},AGGREGATE(15,6,"
        f"ROW({cfg.n})-ROW({cfg.n_anchor})+1/({cfg.n}<>\"\")/(COUNTIF({prev},{cfg.n})=0),1)),\"\")"
    )


def couv(cfg, col, z):
    if col == "C":
        x = cfg.fx
    elif col == "D" and cfg.has_fy:
        x = cfg.fy
    else:
        x = cfg.fz
    inner = (
        f'IFERROR(IFERROR('
        f'LOOKUP(2,1/(({cfg.n}={z})*(LEFT({cfg.b},2)="G ")),{x}),'
        f'LOOKUP(2,1/(({cfg.n}={z})*(ISNUMBER(SEARCH("Couverture",{cfg.b})))),{x})),"-")'
    )
    return wrap(z, inner)


def fx_def(cfg, k, z):
    p = P[k]
    cond = family_has_cases(cfg, p, z)
    inner = (
        f'IFERROR(IF(MINIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z})>=0,'
        f'MAXIFS({cfg.fx},{cfg.b},{p},{cfg.n},{z}),'
        f'INDEX({cfg.fx},MATCH(MINIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z}),{cfg.fz},0))),"-")'
    )
    return wrap_visible(z, cond, inner)


def fy_def(cfg, k, z):
    if not cfg.has_fy:
        return '=""'
    p = P[k]
    cond = family_has_cases(cfg, p, z)
    inner = (
        f'IFERROR(IF(MINIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z})>=0,'
        f'INDEX({cfg.fy},MATCH(MAXIFS({cfg.fx},{cfg.b},{p},{cfg.n},{z}),{cfg.fx},0)),'
        f'INDEX({cfg.fy},MATCH(MINIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z}),{cfg.fz},0))),"-")'
    )
    return wrap_visible(z, cond, inner)


def fz_def(cfg, k, z):
    p = P[k]
    cond = family_has_cases(cfg, p, z)
    inner = (
        f'IFERROR(IF(MINIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z})>=0,'
        f'0,MINIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z})),"-")'
    )
    return wrap_visible(z, cond, inner)


def fx_fav(cfg, k, z):
    p = P[k]
    cond = family_has_uplift(cfg, p, z)
    inner = (
        f'IFERROR(INDEX({cfg.fx},MATCH(MAXIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z}),{cfg.fz},0)),"-")'
    )
    return wrap_visible(z, cond, inner)


def fy_fav(cfg, k, z):
    if not cfg.has_fy:
        return '=""'
    p = P[k]
    cond = family_has_uplift(cfg, p, z)
    inner = (
        f'IFERROR(INDEX({cfg.fy},MATCH(MAXIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z}),{cfg.fz},0)),"-")'
    )
    return wrap_visible(z, cond, inner)


def fz_fav(cfg, k, z):
    p = P[k]
    cond = family_has_uplift(cfg, p, z)
    inner = f'IFERROR(MAXIFS({cfg.fz},{cfg.b},{p},{cfg.n},{z}),"-")'
    return wrap_visible(z, cond, inner)


def neige_n(cfg, col, z):
    if col == "C":
        x = cfg.fx
    elif col == "D" and cfg.has_fy:
        x = cfg.fy
    else:
        x = cfg.fz
    inner = (
        f'IFERROR(INDEX({x},MATCH('
        f'MINIFS({cfg.fz},{cfg.b},"*Neige*",{cfg.b},"<>*accidentel*",{cfg.n},{z}),{cfg.fz},0)),"-")'
    )
    return wrap(z, inner)


def neige_a(cfg, col, z):
    if col == "C":
        x = cfg.fx
    elif col == "D" and cfg.has_fy:
        x = cfg.fy
    else:
        x = cfg.fz
    inner = (
        f'IFERROR(LOOKUP(2,1/(({cfg.n}={z})*(ISNUMBER(SEARCH("accidentel",{cfg.b})))),{x}),"-")'
    )
    return wrap(z, inner)


def row_visible(cfg, mode, z, key=None):
    if mode == "xf":
        return family_has_uplift(cfg, P[key], z)
    if mode == "xd":
        return family_has_cases(cfg, P[key], z)
    return "TRUE"


def formula(cfg, mode, z, key=None):
    if mode == "c":
        if cfg.has_fy:
            return couv(cfg, "C", z), couv(cfg, "D", z), couv(cfg, "E", z)
        return couv(cfg, "C", z), couv(cfg, "E", z)
    if mode == "xd":
        if cfg.has_fy:
            return fx_def(cfg, key, z), fy_def(cfg, key, z), fz_def(cfg, key, z)
        return fx_def(cfg, key, z), fz_def(cfg, key, z)
    if mode == "xf":
        if cfg.has_fy:
            return fx_fav(cfg, key, z), fy_fav(cfg, key, z), fz_fav(cfg, key, z)
        return fx_fav(cfg, key, z), fz_fav(cfg, key, z)
    if mode == "n":
        if cfg.has_fy:
            return neige_n(cfg, "C", z), neige_n(cfg, "D", z), neige_n(cfg, "E", z)
        return neige_n(cfg, "C", z), neige_n(cfg, "E", z)
    if cfg.has_fy:
        return neige_a(cfg, "C", z), neige_a(cfg, "D", z), neige_a(cfg, "E", z)
    return neige_a(cfg, "C", z), neige_a(cfg, "E", z)


def col_at(cfg, offset):
    return cfg.syn_col + offset - 1


def write_saisie_side(ws, cfg, exemple=()):
    title_col = ord(cfg.cas_col) - 64
    last_col = ord(cfg.table_last_col) - 64
    ws.cell(SYN_TITLE, title_col, f"TABLEAU — {cfg.name}").font = SF
    ws.merge_cells(
        start_row=SYN_TITLE, start_column=title_col,
        end_row=SYN_TITLE, end_column=last_col,
    )

    headers = ["Noeud_Cas", "Nom_cas", "FX_daN"]
    if cfg.has_fy:
        headers.extend(["FY_daN", "FZ_daN", "Noeud"])
    else:
        headers.extend(["FZ_daN", "Noeud"])
    for i, h in enumerate(headers):
        col = title_col + i
        c = ws.cell(cfg.table_hdr_row, col, h)
        c.font, c.fill, c.alignment = HF, HDR, CE

    node_idx = ord(cfg.node_col) - 64
    for r in range(DATA_START, DATA_END + 1):
        ws.cell(r, node_idx, node_extract_formula(r, cfg.cas_col))
        for c in range(title_col, last_col + 1):
            ws.cell(r, c).font = BF

    for i, row in enumerate(exemple):
        r = DATA_START + i
        for j, v in enumerate(row):
            cell = ws.cell(r, title_col + j, v)
            cell.font = BF
            cell.fill = INP

    tab_name = "CasChargesFY" if cfg.has_fy else "CasChargesNoFY"
    tab = Table(
        displayName=tab_name,
        ref=f"{cfg.cas_col}{cfg.table_hdr_row}:{cfg.table_last_col}{DATA_END}",
    )
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def build_synthesis_side(ws, cfg):
    title_col = cfg.syn_col
    width = cfg.syn_width
    last_col = title_col + width - 1
    node_idx = ord(cfg.syn_node_col) - 64

    ws.cell(SYN_TITLE, title_col, f"SYNTHESE — {cfg.name}").font = SF
    ws.merge_cells(
        start_row=SYN_TITLE, start_column=title_col,
        end_row=SYN_TITLE, end_column=last_col,
    )

    headers = ["Nom du cas", "FX [daN]"]
    if cfg.has_fy:
        headers.extend(["FY [daN]", "FZ [daN]", "Note"])
    else:
        headers.extend(["FZ [daN]", "Note"])

    for i, h in enumerate(headers):
        c = ws.cell(SYN_HDR, title_col + i, h)
        c.font, c.fill, c.alignment = HF, HDR, CE

    val_offsets = [1, 2, 3] if cfg.has_fy else [1, 2]
    note_offset = 4 if cfg.has_fy else 3

    for block in range(MAX_NODES):
        header_row = SYN_DATA + block * BLOCK
        z = f"${cfg.syn_node_col}${header_row}"
        ws.cell(header_row, node_idx, node_formula(cfg, header_row, cfg.syn_node_col))

        ws.cell(header_row, title_col, f'=IF({z}="","","Noeud "&{z})').font = NF
        ws.merge_cells(
            start_row=header_row, start_column=title_col,
            end_row=header_row, end_column=last_col,
        )
        for c in range(title_col, last_col + 1):
            hcell = ws.cell(header_row, c)
            hcell.fill = PatternFill("solid", fgColor="D9E1F2")
            hcell.border = BD
            hcell.font = NF

        for i, row in enumerate(ROWS):
            r = header_row + 1 + i
            label, mode = row[0], row[1]
            key = row[2] if len(row) > 2 else None
            vis = row_visible(cfg, mode, z, key) if key else "TRUE"
            ws.cell(r, title_col, f'=IF(OR({z}="",NOT({vis})),"","{label}")').font = BF

            vals = formula(cfg, mode, z, key)
            for j, off in enumerate(val_offsets):
                ws.cell(r, title_col + off, vals[j] if j < len(vals) else '=""')
            ws.cell(r, title_col + note_offset, note_dynamic(r, z, cfg.syn_fz_col))

            for c in range(title_col, last_col + 1):
                cell = ws.cell(r, c)
                cell.border = BD
                cell.font = BF
                if c > title_col:
                    cell.alignment = CE


def recalc_excel(path):
    try:
        import win32com.client  # type: ignore

        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible = False
        xl.DisplayAlerts = False
        xl.AskToUpdateLinks = False
        wb = xl.Workbooks.Open(path)
        sa = wb.Worksheets("SAISIE")
        sy = wb.Worksheets("SYNTHESE")

        for cfg in (CFG_FY, CFG_NOFY):
            ncol = ord(cfg.node_col) - 64
            for r in range(DATA_START, DATA_END + 1):
                cell = sa.Cells(r, ncol)
                f = cell.Formula
                if f and str(f).startswith("="):
                    cell.Formula = f

        node_cols = [ord(c.syn_node_col) - 64 for c in (CFG_FY, CFG_NOFY)]
        for r in range(SYN_DATA, SYN_LAST + 1):
            if (r - SYN_DATA) % BLOCK == 0:
                for c in node_cols:
                    cell = sy.Cells(r, c)
                    f = cell.Formula
                    if f and str(f).startswith("="):
                        cell.Formula = f
        xl.CalculateFullRebuild()

        refresh_cols = list(range(1, 6)) + list(range(9, 13)) + node_cols
        for r in range(1, SYN_LAST + 1):
            for c in refresh_cols:
                cell = sy.Cells(r, c)
                f = cell.Formula
                if f and str(f).startswith("="):
                    cell.Formula = f
        xl.CalculateFullRebuild()
        wb.Save()
        wb.Close(False)
        xl.Quit()
        return True
    except Exception:
        return False


def main():
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    me = wb.active
    me.title = "MODE EMPLOI"
    me["A1"] = "DDC Calculateur — 2 tableaux côte à côte"
    me["A1"].font = TF
    for i, t in enumerate(
        [
            "",
            "SAISIE et SYNTHESE : 2 tableaux côte à côte, séparés par les colonnes G-H.",
            "",
            "  GAUCHE (A-F)  : AVEC FY — Noeud_Cas, Nom_cas, FX, FY, FZ",
            "  DROITE (I-M)  : SANS FY — Noeud_Cas, Nom_cas, FX, FZ",
            "",
            "Collez vos données dans UN seul côté (pas les deux).",
            "Note : FZ > 0 = soulèvement, FZ < 0 = compression. F9 si besoin.",
        ],
        3,
    ):
        me.cell(i, 1, t).font = BF
    me.column_dimensions["A"].width = 78

    sa = wb.create_sheet("SAISIE")
    sa["A1"] = "SAISIE — Collez vos cas de charge"
    sa.merge_cells("A1:M1")
    sa["A2"] = "Gauche = avec FY  |  Droite = sans FY  |  Colonnes G-H = espace"
    sa.merge_cells("A2:M2")
    sa["A1"].font = TF
    sa["A2"].font = BF

    write_saisie_side(sa, CFG_FY, EXEMPLE_FY)
    write_saisie_side(sa, CFG_NOFY)

    for col, w in {
        "A": 12, "B": 40, "C": 11, "D": 11, "E": 11, "F": 8,
        "G": 2, "H": 2,
        "I": 12, "J": 40, "K": 11, "L": 11, "M": 8,
    }.items():
        sa.column_dimensions[col].width = w
    sa.freeze_panes = "A5"

    sy = wb.create_sheet("SYNTHESE")
    sy["A1"] = "SYNTHESE — Calcul automatique"
    sy.merge_cells("A1:M1")
    sy["A1"].font = TF
    sy["A2"] = (
        f'="Avec FY : "&COUNTA(SAISIE!$A${DATA_START}:$A${DATA_END})&" cas | "&'
        f'SUMPRODUCT(--(MOD(ROW($F${SYN_DATA}:$F${SYN_LAST})-{SYN_DATA},{BLOCK})=0)*($F${SYN_DATA}:$F${SYN_LAST}<>""))&" noeud(s)  ||  "&'
        f'"Sans FY : "&COUNTA(SAISIE!$I${DATA_START}:$I${DATA_END})&" cas | "&'
        f'SUMPRODUCT(--(MOD(ROW($M${SYN_DATA}:$M${SYN_LAST})-{SYN_DATA},{BLOCK})=0)*($M${SYN_DATA}:$M${SYN_LAST}<>""))&" noeud(s)"'
    )
    sy["A2"].font = BF
    sy.merge_cells("A2:M2")

    build_synthesis_side(sy, CFG_FY)
    build_synthesis_side(sy, CFG_NOFY)

    sy.column_dimensions["F"].hidden = True
    sy.column_dimensions["M"].hidden = True
    for col, w in {
        "A": 22, "B": 12, "C": 12, "D": 12, "E": 14,
        "G": 2, "H": 2,
        "I": 22, "J": 12, "K": 12, "L": 14,
    }.items():
        sy.column_dimensions[col].width = w
    sy.freeze_panes = "A5"

    out = OUTPUT
    try:
        wb.save(out)
    except PermissionError:
        out = OUTPUT.replace(".xlsx", " (new).xlsx")
        wb.save(out)

    print("Fichier:", out)
    if recalc_excel(out):
        print("Recalcul Excel OK — valeurs visibles a l'ouverture.")
    else:
        print("Ouvrez dans Excel et appuyez sur F9 si les cellules sont vides.")


if __name__ == "__main__":
    main()